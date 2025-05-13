from rest_framework import viewsets, generics, status
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from .serializers import *
from .tasks import auto_submit_exam
from users.serializers import UserSerializer
from .models import *
from subscription.models import *
from .permissions import IsAdminOrReadOnly, IsAdmin
from django.utils import timezone
from rest_framework.pagination import PageNumberPagination
from django.shortcuts import get_object_or_404
from django.views.generic.detail import DetailView
from rest_framework_simplejwt.authentication import JWTAuthentication
from django.utils.dateparse import parse_date
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models.functions import ExtractMonth, ExtractYear
from django.db import transaction, IntegrityError
from django.contrib.auth.decorators import login_required
from rest_framework.decorators import api_view, permission_classes
from rest_framework.exceptions import ValidationError
from datetime import timedelta
from django.db.models import Q, Count, Sum
from random import sample
from invitation.models import ExamInvite
from calendar import monthrange
from django.http import JsonResponse
import openpyxl
import tempfile
import random
import json
import uuid
from django.core.exceptions import ObjectDoesNotExist
# import logging
import pandas as pd
from .pagination import CustomPageNumberPagination
from django.contrib.auth import get_user_model
from django.core.exceptions import PermissionDenied
from django.db.models import F, Window
from django.db.models.functions import Rank
User = get_user_model()
import os
import traceback
now = timezone.now()
import logging
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile

# Configure logging at the top of your file
logging.basicConfig(level=logging.INFO)



class CategoryListView(APIView):
    def get(self, request):
        categories = ExamCategory.objects.all()
        serializer = ExamCategorySerializer(categories, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

# View to create a new category
class CreateCategoryView(APIView):
    def post(self, request):
        serializer = ExamCategorySerializer(data=request.data)
        if serializer.is_valid():
            # Check if category with the same name exists
            if ExamCategory.objects.filter(name=serializer.validated_data['name']).exists():
                return Response({"error": "Category already exists."}, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)





class LeaderboardListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, exam_id):
        try:
            leaderboards = Leaderboard.objects.filter(exam_id=exam_id).order_by('-score')[:10]  # Get top 10 scores for the exam
            serializer = LeaderboardSerializer(leaderboards, many=True)
            # print(serializer.data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Leaderboard.DoesNotExist:
            raise Http404("Leaderboard does not exist for this exam.")


class ExamDetailView(generics.RetrieveAPIView):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]
    lookup_field = 'exam_id'

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"request": self.request})
        return context
    
    
    def get(self, request, *args, **kwargs):
        # Retrieve the exam instance
        exam = self.get_object()

        # Check if the user can access the exam
        

        
        # Proceed with the default behavior if access is allowed
        return super().get(request, *args, **kwargs)


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer



    

class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.all()
    serializer_class = ExamSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes=[JWTAuthentication]

    
    def perform_create(self, serializer):
        # print('Hello')
        # Check if the 'last_date' is provided in the validated data
        # last_date = self.request.data.get('last_date', None)
        # print("hello");
        # Save the exam with the 'created_by' and 'last_date' if provided
        if serializer.is_valid():
            exam = serializer.save(created_by=self.request.user)
        else:
            print(serializer.errors)
        status = 'draft'

        if self.request.user.role == 'student':
            status = 'student'
            Status.objects.create(exam=exam, status=status, user=self.request.user)
            return Response({'exam_id': exam.exam_id})

        # Create the status for the exam
        Status.objects.create(exam=exam, status=status, user=self.request.user)
        # Leaderboard.objects.create(exam=exam, score=0)
        return Response({'exam_id': exam.exam_id})
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated], authentication_classes=[JWTAuthentication])
    def exam_list(self, request):
        """
        Get exams filtered by difficulty ranges based on the user's subscription package.
        """
        # Fetch the active subscription for the user
        active_subscription = UserSubscription.objects.filter(user=request.user, status='active', end_date__gte=now.date()).first()
        
        if not active_subscription or not active_subscription.package:
            return Response({"error": "You do not have an active subscription package."}, status=400)

        # Retrieve the subscription package
        subscription_package = active_subscription.package

        # Retrieve exams with published status
        exams = Exam.objects.filter(exam__status='published')
        filtered_exams = []

        def parse_range(range_str):
            """
            Helper function to parse range strings like "10-20" into min and max integers.
            """
            try:
                range_parts = range_str.split('-')
                if len(range_parts) != 2:
                    raise ValueError
                return int(range_parts[0].strip()), int(range_parts[1].strip())
            except ValueError:
                raise ValueError(f"Invalid range format: {range_str}")

        try:
            # Parse range values from the subscription package
            very_easy_min, very_easy_max = parse_range(subscription_package.very_easy_percentage)
            easy_min, easy_max = parse_range(subscription_package.easy_percentage)
            medium_min, medium_max = parse_range(subscription_package.medium_percentage)
            hard_min, hard_max = parse_range(subscription_package.hard_percentage)
            very_hard_min, very_hard_max = parse_range(subscription_package.very_hard_percentage)
            expert_min, expert_max = parse_range(subscription_package.expert_percentage)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)

        for exam in exams:
            # Ensure the exam has a related difficulty instance
            difficulty = getattr(exam, 'difficulty', None)
            if not difficulty:
                continue

            # Validate difficulty percentages against subscription ranges
            if (
                very_easy_min <= difficulty.difficulty1_percentage <= very_easy_max
                and easy_min <= difficulty.difficulty2_percentage <= easy_max
                and medium_min <= difficulty.difficulty3_percentage <= medium_max
                and hard_min <= difficulty.difficulty4_percentage <= hard_max
                and very_hard_min <= difficulty.difficulty5_percentage <= very_hard_max
                and expert_min <= difficulty.difficulty6_percentage <= expert_max
            ):
                filtered_exams.append(exam)
        
        
        # Serialize the filtered exams
        serializer = self.get_serializer(filtered_exams, many=True)
        return Response(serializer.data)
    
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated], authentication_classes=[JWTAuthentication])
    def student_exam_list(self, request):
        
        exams = Exam.objects.filter(exam__status='student', created_by=self.request.user)
        serializer = self.get_serializer(exams, many=True)
        return Response(serializer.data)
    
    
    
    @action(detail=True, methods=['get'], url_path='start', permission_classes=[IsAuthenticated])
    def start_exam(self, request, pk=None):
        # Fetch the exam instance
        exam = self.get_object()

        # Retrieve the exam's status
        status_instance = Status.objects.filter(exam=exam).first()
        # print(status_instance)
        # Fetch the user's usage tracking and subscription package
        usage_tracking = UsageTracking.objects.filter(user=request.user).first()
        
        
        
        if not usage_tracking or not usage_tracking.package:
            return JsonResponse({"error": "You do not have an active subscription package."}, status=403)

        # Check if the user can take the exam
        if not usage_tracking.can_take_exam(exam.exam_id):
            return JsonResponse({"error": "Exam limit exceeded for this subscription."}, status=403)

        # Check if the user can attempt this specific exam
        if not usage_tracking.can_attempt_exam(exam.exam_id):
            return JsonResponse({"error": "Maximum attempts for this exam exceeded."}, status=403)

        # Validate exam status
        if status_instance and status_instance.status in ['published', 'student']:
            # Record the exam start time
            start_time = now

            # Optional: Calculate end time if `duration` exists
            end_time = start_time + (exam.duration or timezone.timedelta())

            # Increment usage and attempts
            try:
                
                # Increment attempt count for this specific exam
                exam_attempts = usage_tracking.exam_attempts
                # exam_attempts[str(exam.exam_id)] = exam_attempts.get(str(exam.exam_id), {"attempts": 0})
                # exam_attempts[str(exam.exam_id)]["attempts"] += 1
                usage_tracking.exam_attempts = exam_attempts
                # usage_tracking.total_attempts_taken += 1 
                usage_tracking.save()
            except ValueError as e:
                return JsonResponse({"error": str(e)}, status=403)

            return Response({
                'exam_id': exam.exam_id,
                'title': exam.title,
                'total_questions': exam.total_questions,
                'duration': exam.duration,
                'start_time': start_time,
                'end_time': end_time,
            })

        return Response({"error": "The exam is not published."}, status=status.HTTP_403_FORBIDDEN)


    @action(detail=True, methods=['get'], url_path='questions', permission_classes=[IsAuthenticated])
    def get_questions(self, request, pk=None):
        exam = self.get_object()

        # Fetch the user's usage tracking
        usage_tracking = UsageTracking.objects.filter(user=request.user).first()
        if not usage_tracking or not usage_tracking.package:
            return JsonResponse({"error": "You do not have an active subscription package."}, status=403)


        if not usage_tracking.can_take_exam(pk):
            return JsonResponse({"error": "Exam limit exceeded for this subscription."}, status=403)

        # Check if the user has started the exam
        # if not usage_tracking.exam_attempts.get(str(exam.exam_id)):
        #     return JsonResponse({"error": "You have not started this exam yet."}, status=403)

        usage_tracking.increment_exam(pk)
        # if usage_tracking.total:
        #     return JsonResponse({"error": "Exam limit exceeded for this subscription."}, status=403)
        
        # Fetch and serialize the questions
        questions = exam.questions.all()
        serializer = QuestionSerializer(questions, many=True)

        return Response({
            "questions": serializer.data,
            "skipped_questions": []  # Initially empty, will hold skipped question IDs as user skips questions
        })

    

    @action(detail=True, methods=['post'], url_path='submit', permission_classes=[IsAuthenticated])
    def submit_exam(self, request, pk=None):
        user = request.user
        exam = self.get_object()
        usage_tracking = UsageTracking.objects.filter(user=request.user).first()
        
        package = usage_tracking.package
        
        if not usage_tracking.exam_attempts.get(str(pk)):
            return JsonResponse({"error": "You have not started this exam yet."}, status=403)
        
        max_attempts = package.max_attampts
        if usage_tracking.exam_attempts[str(pk)]["attempts"] >= max_attempts+1:
            return JsonResponse({"error":"You need to update your package"})
        
        if usage_tracking.total_exams_taken >= package.max_exams+1:
            return JsonResponse({"error":"You need to update your package"})
        
        # Create a new attempt
        attempt = ExamAttempt.objects.create(
            exam=exam,
            user=user,
            answered=0,
            total_correct_answers=0,
            wrong_answers=0,
            passed=False
        )

        # Retrieve answers from the request
        answers = request.data.get('answers', [])
        if not isinstance(answers, list):
            return Response({"error": "Answers should be a list."}, status=status.HTTP_400_BAD_REQUEST)

        # Initialize counters
        total_correct = 0
        total_wrong = 0
        answered_count = 0

        # Process each answer
        for answer in answers:
            if answer is None:
                continue  # Skip if answer is None

            question_id = answer.get('question_id')
            selected_option_id = answer.get('option')

            if question_id is None or selected_option_id == 'none':
                continue  # Skip if either question_id or selected_option is missing

            try:
                question = Question.objects.get(id=question_id)
                selected_option = QuestionOption.objects.get(id=int(selected_option_id))
            except (Question.DoesNotExist, QuestionOption.DoesNotExist, ValueError):
                raise ValidationError({"error": "Invalid question or option ID."})

            correct_answer = question.get_correct_answer()
            is_correct = selected_option.is_correct

            # Count as answered
            answered_count += 1

            if is_correct:
                total_correct += 1
            else:
                total_wrong += 1

            # Create or update UserAnswer entry for the question and attempt
            UserAnswer.objects.update_or_create(
                exam_attempt=attempt,
                question=question,
                defaults={
                    'selected_option': selected_option,
                    'is_correct': is_correct
                }
            )

        # Update the attempt with results
        attempt.answered = answered_count
        attempt.total_correct_answers = total_correct
        attempt.wrong_answers = total_wrong
        attempt.passed = total_correct >= exam.pass_mark  # Adjust pass logic as needed
        attempt.save()

        Leaderboard.update_best_score(user, exam)
        # Schedule auto-submit task (if needed)
        auto_submit_exam.apply_async(args=[attempt.id], countdown=10)

        return JsonResponse({
            'status': 'submitted',
            'answered': answered_count,
            'correct_answers': total_correct,
            'wrong_answers': total_wrong,
            'passed': attempt.passed
        }, status=status.HTTP_200_OK)
        
        
    @action(detail=True, methods=['post'], url_path='skip', permission_classes=[IsAuthenticated])
    def skip_question(self, request, pk=None):
        exam = self.get_object()
        skipped_question_id = request.data.get('question_id')

        if not skipped_question_id:
            return Response({"error": "Question ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Add skipped question ID to session or cache for the current user-exam attempt
        skipped_questions = request.session.get(f"skipped_questions_{exam.exam_id}", [])
        
        if skipped_question_id not in skipped_questions:
            skipped_questions.append(skipped_question_id)
            request.session[f"skipped_questions_{exam.exam_id}"] = skipped_questions

        return Response({"message": f"Question {skipped_question_id} skipped."})
    
    @action(detail=True, methods=['post'])
    def generate_exam(self, request, pk=None):
        """
        Custom action to generate an exam by selecting random questions based on difficulty percentages.
        """
        exam_id = pk  # Get the exam ID from the URL
        
        # Ensure total_questions is treated as an integer
        try:
            total_questions = int(request.data.get('total_questions', 10))  # Cast to int
        except ValueError:
            return Response({'error': 'Invalid total questions number.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Fetch the exam and its associated difficulty percentages
            exam = Exam.objects.get(exam_id=exam_id)
            total_questions = exam.total_questions
            difficulty = ExamDifficulty.objects.get(exam=exam)
        except Exam.DoesNotExist:
            return Response({'error': 'Exam not found.'}, status=status.HTTP_404_NOT_FOUND)
        except ExamDifficulty.DoesNotExist:
            return Response({'error': 'Difficulty settings not found for this exam.'}, status=status.HTTP_404_NOT_FOUND)

        # Retrieve difficulty percentages
        difficulty_percentages = {
            'difficulty1': difficulty.difficulty1_percentage,
            'difficulty2': difficulty.difficulty2_percentage,
            'difficulty3': difficulty.difficulty3_percentage,
            'difficulty4': difficulty.difficulty4_percentage,
            'difficulty5': difficulty.difficulty5_percentage,
            'difficulty6': difficulty.difficulty6_percentage,
        }

        # Calculate the number of questions for each difficulty level based on the total question count
        question_distribution = {
            'difficulty1': round((difficulty_percentages['difficulty1'] / 100) * total_questions),
            'difficulty2': round((difficulty_percentages['difficulty2'] / 100) * total_questions),
            'difficulty3': round((difficulty_percentages['difficulty3'] / 100) * total_questions),
            'difficulty4': round((difficulty_percentages['difficulty4'] / 100) * total_questions),
            'difficulty5': round((difficulty_percentages['difficulty5'] / 100) * total_questions),
            'difficulty6': total_questions - (
                round((difficulty_percentages['difficulty1'] / 100) * total_questions) +
                round((difficulty_percentages['difficulty2'] / 100) * total_questions) +
                round((difficulty_percentages['difficulty3'] / 100) * total_questions) +
                round((difficulty_percentages['difficulty4'] / 100) * total_questions) +
                round((difficulty_percentages['difficulty5'] / 100) * total_questions)
            )
        }

        # Fetch published questions for each difficulty level
        questions_by_difficulty = {
            'difficulty1': list(Question.objects.filter(status='published', difficulty_level=1)),
            'difficulty2': list(Question.objects.filter(status='published', difficulty_level=2)),
            'difficulty3': list(Question.objects.filter(status='published', difficulty_level=3)),
            'difficulty4': list(Question.objects.filter(status='published', difficulty_level=4)),
            'difficulty5': list(Question.objects.filter(status='published', difficulty_level=5)),
            'difficulty6': list(Question.objects.filter(status='published', difficulty_level=6)),
        }

        # Randomly sample the required number of questions for each difficulty
        selected_questions = []
        for difficulty_level, question_count in question_distribution.items():
            questions = questions_by_difficulty[difficulty_level]
            if question_count > 0:
                selected_questions += sample(questions, min(len(questions), question_count))

        print(len(selected_questions))
        # Check if we have enough questions selected
        if len(selected_questions) < total_questions:
            return Response({'error': 'Not enough questions available to generate the exam.'}, status=status.HTTP_400_BAD_REQUEST)

        # Associate the selected questions with the exam
        exam.questions.set(selected_questions)
        exam.save()

        return Response({
            'message': 'Exam generated successfully.',
            'exam_id': exam.exam_id,
            'questions': [q.id for q in selected_questions]
        }, status=status.HTTP_200_OK)

    #admin attampt check
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def all_exams_with_attempts(self, request):
        # Get all exams and annotate them with the count of related ExamAttempt objects
        exams = Exam.objects.annotate(num_attempts=Count('attempts'))
        
        # Prepare the response data
        data = []
        for exam in exams:
            data.append({
                'exam_id': str(exam.exam_id),  # Assuming exam.id is UUID
                'exam_title': exam.title,
                'num_attempts': exam.num_attempts
            })

        return Response(data)
    
    @action(detail=True, methods=['get'], permission_classes=[IsAuthenticated])
    def user_attempts(self, request, pk=None):
        print(pk)
        # Get the exam by its primary key (UUID)
        try:
            exam = Exam.objects.get(exam_id=pk)
        except Exam.DoesNotExist:
            return Response({"error": "Exam not found."}, status=404)

        # Fetch users who attempted the specific exam and count their attempts
        user_attempts = (
            ExamAttempt.objects.filter(exam=exam)
            .values('user__username')
            .annotate(num_attempts=Count('id'))  # Count the number of attempts per user
            .order_by('-num_attempts')
        )

        # Prepare the response data
        data = []
        for user_attempt in user_attempts:
            data.append({
                'username': user_attempt['user__username'],
                'num_attempts': user_attempt['num_attempts'],
            })

        return Response(data)
    
    

class ExamDifficultyView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = ExamDifficultySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Exam difficulty added successfully"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    
class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]

    def perform_create(self, serializer):
        exam = serializer.validated_data.get('exam')
        question = serializer.save(exam=exam)
        return Response({"question_id": question.id})

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated, IsAdminOrReadOnly])
    def add_option(self, request, pk=None):
        question = self.get_object()
        serializer = QuestionOptionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(question=question)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def upload_questions(self, request):
        if 'file' not in request.FILES:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if 'Exam', 'Year', and 'Subject' are in the request body
        exam_name = request.data.get('exam_name')
        exam_year = request.data.get('exam_year')
        # subject_name = request.data.get('Subject')
        print(exam_name, exam_year)

        if not exam_name or not exam_year:
            return Response({"error": "Exam name and year must be provided in the request body."}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']

        try:
            df = pd.read_excel(file)
        except Exception as e:
            return Response({"error": f"Error reading file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        required_columns = ['Question', 'Option1', 'Option2', 'Option3', 'Option4', 'Answer', 'Options_num', 'Category', 'Difficulty', 'Subject']
        if not all(col in df.columns for col in required_columns):
            missing_cols = [col for col in required_columns if col not in df.columns]
            return Response({"error": f"Missing columns: {', '.join(missing_cols)}"}, status=status.HTTP_400_BAD_REQUEST)

        question_count = 0
        duplicate_questions = []
        error_details = []

        try:
            with transaction.atomic():
                

                for _, row in df.iterrows():
                    subject_name = row['Subject']
                    subject, created = Subject.objects.get_or_create(name=subject_name)
                    
                    question_text = row['Question']
                    options = [row['Option1'], row['Option2'], row['Option3'], row['Option4']]
                    correct_answer = row['Answer'].strip().lower().replace(" ", "")
                    option_num = row['Options_num']
                    category_name = row['Category']
                    difficulty_level = int(row['Difficulty'])

                    if difficulty_level not in range(1, 7):
                        error_details.append(f"Question '{question_text}' has invalid difficulty level {difficulty_level}. It must be between 1 and 6.")
                        continue

                    # Get or create category
                    category, created = Category.objects.get_or_create(name=category_name)

                    # Create or retrieve the question
                    question, created = Question.objects.get_or_create(
                        text=question_text,
                        defaults={
                            'marks': 1,
                            'category': category,
                            'subject': subject,
                            'difficulty_level': difficulty_level,
                            'created_by': request.user,
                            'status': 'submitted'
                        }
                    )

                    if created:
                        question_count += 1

                        # Create the options for the new question
                        for i, option_text in enumerate(options, start=1):
                            normalized_option_label = f"option{i}".strip().lower().replace(" ", "")
                            is_correct = (normalized_option_label == correct_answer)
                            QuestionOption.objects.create(
                                question=question,
                                text=option_text,
                                is_correct=is_correct
                            )
                    else:
                        duplicate_questions.append(question_text)

                    # Check for and update QuestionUsage for this exam and year
                    usage_exists = QuestionUsage.objects.filter(
                        question=question,
                        exam=exam_name,
                        year=exam_year
                    ).exists()

                    if not usage_exists:
                        QuestionUsage.objects.create(
                            question=question,
                            exam=exam_name,
                            year=exam_year
                        )

        except Exception as e:
            return Response({"error": f"Error while saving questions: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        response_message = f"{question_count} questions uploaded successfully."
        if duplicate_questions:
            response_message += f" Duplicate questions skipped: {', '.join(duplicate_questions)}"
        if error_details:
            response_message += f" Errors: {', '.join(error_details)}"

        return Response({"message": response_message}, status=status.HTTP_200_OK)



    
    @action(detail=False, methods=['get'], permission_classes=[IsAdminOrReadOnly])
    def user_questions(self, request):
        user = self.request.user
        
        user_questions = Question.objects.filter(created_by=user)  # Filter by the logged-in user
        # print(user_questions)
        serializer = self.get_serializer(user_questions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def submitted_users(self, request):
        # Get users who have submitted questions (status='submitted')
        submitted_status = 'submitted'
        users = User.objects.filter(
            question_created_by__status=submitted_status
        ).annotate(total_submitted_questions=Count('question_created_by')).distinct()

        # Prepare the user data for the response
        users_data = [
            {
                'username': user.username,
                'total_questions': user.total_submitted_questions,
                'user_id': user.id
            }
            for user in users
        ]

        # Return the list of users with their total submitted questions
        return Response(users_data, status=status.HTTP_200_OK)

    # View all submitted questions of a specific user
    @action(detail=False, methods=['get'], url_path='submitted_questions/<int:user_id>')
    def submitted_questions(self, request, user_id=None):
        submitted_status = 'submitted'
        user = User.objects.get(id=user_id)
        # print('Hello')

        user_questions = Question.objects.filter(created_by=user, status=submitted_status)  # Filter by the logged-in user
        serializer = self.get_serializer(user_questions, many=True)
        return Response(serializer.data)

    # List users with reviewed questions and show total reviewed questions
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def reviewed_users(self, request):
        """
        Custom view to get users whose questions have a status of 'reviewed' and 
        were reviewed by the currently logged-in user, along with the total number of such questions.
        """
        user = self.request.user

        # Filter questions that are reviewed by the logged-in user and have a status of 'reviewed'
        questions = Question.objects.filter(status='reviewed', reviewed_by=user)

        # Aggregate data to count the total number of questions reviewed by the logged-in user
        users = User.objects.filter(
            question_created_by__in=questions
        ).annotate(total_reviewed_questions=Count('question_created_by')).distinct()

        # Prepare user data with total reviewed questions
        users_data = [
            {
                'username': user.username,
                'total_reviewed_questions': user.total_reviewed_questions,
                'user_id': user.id
            }
            for user in users
        ]

        # Return the list of users with the total number of reviewed questions
        return Response(users_data, status=status.HTTP_200_OK)

    # View all reviewed questions of a specific user
    @action(detail=True, methods=['get'], permission_classes=[IsAdminOrReadOnly])
    def reviewed_questions(self, request, pk=None):
        reviewed_status = 'reviewed'
        # print('hello')
        # print("pk: ", pk)
        # Get the reviewer (the logged-in user)
        reviewer =self.request.user
        # print(reviewer)

        # If user_id is provided, filter questions created by that user and reviewed by the current reviewer
        if pk:
            user = User.objects.get(id=pk)
            # Filter questions created by the user and reviewed by the current reviewer with a "reviewed" status
            questions = Question.objects.filter(
                created_by=user,
                status=reviewed_status,
                reviewed_by=reviewer
            ).distinct()
        else:
            # If no user_id is provided, just get the questions reviewed by the current reviewer
            questions = Question.objects.filter(
                status_history__status=reviewed_status,
                status_history__user=reviewer
            ).distinct()

        # Serialize the reviewed questions
        serializer = self.get_serializer(questions, many=True)

        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def approved_users(self, request):
        """
        Custom view to get users whose questions have a status of 'reviewed' and 
        were reviewed by the currently logged-in user, along with the total number of such questions.
        """
        user = self.request.user

        # Filter questions that are reviewed by the logged-in user and have a status of 'reviewed'
        questions = Question.objects.filter(status='approved')

        # Aggregate data to count the total number of questions reviewed by the logged-in user
        users = User.objects.filter(
            question_created_by__in=questions
        ).annotate(total_reviewed_questions=Count('question_created_by')).distinct()
        # print(users)
        # Prepare user data with total reviewed questions
        users_data = [
            {
                'username': user.username,
                'total_approved_questions': user.total_reviewed_questions,
                'reviewer': questions.filter(created_by=user).first().reviewed_by.username if questions.filter(created_by=user).first() else None,
                'user_id': user.id
            }
            for user in users
        ]

        # Return the list of users with the total number of reviewed questions
        return Response(users_data, status=status.HTTP_200_OK)
    
        
    @action(detail=True, methods=['get'], permission_classes=[IsAdminOrReadOnly])
    def approved_questions(self, request, pk=None):
        reviewed_status = 'approved'
        # print('hello')
        # print("pk: ", pk)
        # Get the reviewer (the logged-in user)
        user = self.request.user

        # If user_id is provided, filter questions created by that user and reviewed by the current reviewer
        if pk:
            user = User.objects.get(id=pk)
            # Filter questions created by the user and reviewed by the current reviewer with a "reviewed" status
            questions = Question.objects.filter(
                created_by=user,
                status=reviewed_status,
            ).distinct()
        else:
            # If no user_id is provided, just get the questions reviewed by the current reviewer
            questions = Question.objects.filter(
                status_history__status=reviewed_status,
                status_history__user=reviewer
            ).distinct()

        # Serialize the reviewed questions
        serializer = self.get_serializer(questions, many=True)
        # id = pk
        # data ={"id": id, "questions": serializer.data}
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def submit_all_reviews(self, request):
        reviews = request.data.get('reviews', [])
        
        if not reviews:
            return Response({"error": "No reviews provided"}, status=status.HTTP_400_BAD_REQUEST)

        question_ids = [review.get('question_id') for review in reviews]
        remarks_by_id = {review.get('question_id'): review.get('remarks', '') for review in reviews}
        
        # val = remarks_by_id.get('1760', '')
        # print(val)
        # print("id: ", question_ids)
        # print("remarks: ", remarks_by_id)

        # Use a transaction to wrap the whole process
        with transaction.atomic():
            questions = Question.objects.filter(id__in=question_ids)
            # print(questions)
            for question in questions:
                # print(remarks_by_id.get(question.id))
                # print("remark", remarks_by_id.get(question.id, ''))
                print(question.id)
                question.remarks = remarks_by_id.get(str(question.id), '')
                question.status = 'approved'
            
            Question.objects.bulk_update(questions, ['remarks', 'status'])

        return Response({"message": "All reviews processed successfully"}, status=status.HTTP_200_OK)
    
    
    
    @action(detail=False, methods=['post'], permission_classes=[IsAdmin])
    def assign_teacher(self, request):
        """
        Custom action to assign a teacher to multiple questions and update their status to reviewed
        """
        teacher_id = self.request.data.get('teacherId')
        question_ids = self.request.data.get('question_ids', [])
        print("teacher id:", teacher_id)
        if not teacher_id:
            return Response({'error': 'Teacher id is required.'}, status=status.HTTP_400_BAD_REQUEST)

        if not question_ids or not isinstance(question_ids, list):
            return Response({'error': 'A list of question IDs is required.'}, status=status.HTTP_400_BAD_REQUEST)

        teacher = get_object_or_404(User, id=teacher_id, role='teacher')

        questions = Question.objects.filter(id__in=question_ids).update(reviewed_by=teacher, status='reviewed')
        updated_count = 0
             
        return Response({'message': f'{updated_count} questions updated.'}, status=status.HTTP_200_OK)
    
    
    @action(detail=True, methods=['post'], permission_classes=[IsAdmin])
    def publish_approved(self, request, pk):
        user = get_object_or_404(User, id=pk)

        # Filter questions created by the user with status 'approved'
        approved_questions = Question.objects.filter(created_by=user, status='approved')

        if not approved_questions.exists():
            return Response({'error': 'No approved questions found for this user.'}, status=status.HTTP_404_NOT_FOUND)

        # Split questions based on whether they have remarks
        questions_with_remarks = approved_questions.filter(~Q(remarks=''))  # Questions with remarks
        questions_without_remarks = approved_questions.filter(remarks='')  # Questions without remarks

        # Update the status of questions with remarks to 'rejected'
        rejected_count = questions_with_remarks.update(status='rejected')

        # Update the status of questions without remarks to 'published'
        published_count = questions_without_remarks.update(status='published')

        return Response({
            'message': f'{published_count} questions published successfully and {rejected_count} questions rejected.'
        }, status=status.HTTP_200_OK)


    @action(detail=False, methods=['get'], permission_classes=[IsAdminOrReadOnly])
    def question_bank(self, request):
        try:
            # Get all questions with 'approved' status
            questions = Question.objects.filter(status='published')
            paginator = CustomPageNumberPagination()
            paginated_questions = paginator.paginate_queryset(questions, request)
            
            # Serialize the updated questions
            serializer = self.get_serializer(paginated_questions, many=True)

            return paginator.get_paginated_response(serializer.data)

        except Exception as e:
            # Log the error message
            print(f"Error in question_bank view: {e}")
            return Response({'error': 'An error occurred while fetching questions.'}, status=500)

   
    @action(detail=False, methods=['get'], permission_classes=[IsAdminOrReadOnly])
    def get_remarks(self, request):
        questions = self.get_queryset().exclude(remarks='')
        serializer = self.get_serializer(questions, many=True)
        return Response(serializer.data)
    
class QuestionOptionViewSet(viewsets.ModelViewSet):
    queryset = QuestionOption.objects.all()
    serializer_class = QuestionOptionSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    authentication_classes = [JWTAuthentication]
    
    def perform_create(self, serializer):
        question = serializer.validated_data.get('question')  # Ensure exam is included
        serializer.save(question=question)

    
    
class UserCreatedExamsView(ListAPIView):
    serializer_class = ExamSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]

    def get_queryset(self):
        user = self.request.user
        return Exam.objects.filter(created_by=user)


# class ExamUploadView(APIView):
#     def post(self, request, *args, **kwargs):
#         exam_id = request.POST.get('exam_id')

#         if not exam_id:
#             return Response({"error": "Exam ID is required."}, status=status.HTTP_400_BAD_REQUEST)

#         if 'file' not in request.FILES:
#             return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             exam = Exam.objects.get(exam_id=exam_id)
#         except Exam.DoesNotExist:
#             return Response({"error": "Exam not found."}, status=status.HTTP_404_NOT_FOUND)

#         file = request.FILES['file']
#         try:
#             df = pd.read_excel(file)
#         except Exception as e:
#             return Response({"error": f"Error reading file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

#         required_columns = ['Question', 'Option1', 'Option2', 'Option3', 'Option4', 'Answer', 'Options_num', 'Category', 'Difficulty', 'Subject']
#         df_columns_lower = [col.lower() for col in df.columns]

#         # Check for missing columns in a case-insensitive manner
#         missing_cols = [col for col in required_columns if col.lower() not in df_columns_lower]
#         if missing_cols:
#             return Response({"error": f"Missing columns: {', '.join(missing_cols)}"}, status=status.HTTP_400_BAD_REQUEST)

#         question_count = 0
#         duplicate_questions = []

#         try:
#             with transaction.atomic():
#                 for _, row in df.iterrows():
#                     # Normalize column access to be case-insensitive
#                     row = {col.lower(): value for col, value in row.items()}

#                     question_text = row['question']
#                     options = [row['option1'], row['option2'], row['option3'], row['option4']]
#                     correct_answer = row['answer'].strip().lower().replace(" ", "")

#                     # Extract subject from the row
#                     subject_name = row['subject']
#                     subject, created = Subject.objects.get_or_create(name=subject_name)

#                     # Check if the question already exists for this exam (to avoid duplicates)
#                     if Question.objects.filter(exam=exam, text=question_text).exists():
#                         duplicate_questions.append(question_text)
#                         continue

#                     # Handle the difficulty level and category
#                     category_name = row['category']
#                     try:
#                         difficulty_level = int(row['difficulty'])  # Assuming difficulty is an integer
#                     except ValueError:
#                         return Response({"error": f"Invalid difficulty level '{row['difficulty']}'. It must be an integer."}, status=status.HTTP_400_BAD_REQUEST)

#                     if difficulty_level not in range(1, 7):  # Validate difficulty level
#                         return Response({"error": f"Invalid difficulty level {difficulty_level}. It must be between 1 and 6."}, status=status.HTTP_400_BAD_REQUEST)

#                     category, created = Category.objects.get_or_create(name=category_name)
#                     question_count += 1

#                     # Create the question object
#                     question = Question.objects.create(
#                         exam=exam,
#                         text=question_text,
#                         marks=1,  # Assuming static marks for each question
#                         category=category,
#                         difficulty_level=difficulty_level,
#                         subject=subject  # Set the subject
#                     )

#                     # Create the options and mark the correct one
#                     for i, option_text in enumerate(options, start=1):
#                         normalized_option_label = f"option{i}".strip().lower().replace(" ", "")
#                         is_correct = (normalized_option_label == correct_answer)
#                         QuestionOption.objects.create(
#                             question=question,
#                             text=option_text,
#                             is_correct=is_correct
#                         )

#                 # Update the total number of questions in the exam
#                 exam.total_questions = question_count
#                 exam.save()

#                 # Provide feedback on duplicate questions
#                 if duplicate_questions:
#                     return Response({"message": "Some questions were duplicates and skipped.", "duplicates": duplicate_questions}, status=status.HTTP_201_CREATED)
#                 else:
#                     return Response({"message": "All questions created successfully."}, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             return Response({"error": f"An error occurred: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

def save_image_with_unique_name(image_data, original_filename, upload_to='uploads'):
    """Saves image data to a Django ContentFile with a unique filename."""
    try:
        img = Image.open(BytesIO(image_data))
        img_format = img.format or "PNG"
        unique_filename = f"{uuid.uuid4().hex}.{img_format.lower()}"
        img_io = BytesIO()
        img.save(img_io, format=img_format)
        return ContentFile(img_io.getvalue(), name=os.path.join(upload_to, unique_filename))
    except Exception as e:
        logging.error(f"Error saving image: {e}")
        return None

class ExamUploadView(APIView):
    def post(self, request, *args, **kwargs):
        exam_id = request.POST.get('exam_id')

        if not exam_id:
            return Response({"error": "Exam ID is required."}, status=status.HTTP_400_BAD_REQUEST)

        if 'file' not in request.FILES:
            return Response({"error": "No Excel file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            exam = Exam.objects.get(exam_id=exam_id)
        except Exam.DoesNotExist:
            return Response({"error": "Exam not found."}, status=status.HTTP_404_NOT_FOUND)

        excel_file = request.FILES['file']

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            return Response({"error": f"Error reading Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        df_columns_lower = [col.lower() for col in df.columns]

        def get_column_data(col_name, row):
            if col_name.lower() in df_columns_lower:
                return row[df.columns[df_columns_lower.index(col_name.lower())]]
            return None

        required_columns_text = ['Question_Text', 'Option1_Text', 'Option2_Text', 'Option3_Text', 'Option4_Text', 'Answer', 'Options_num', 'Category', 'Difficulty', 'Subject']
        missing_cols_text = [col for col in required_columns_text if col.lower() not in df_columns_lower]
        if missing_cols_text:
            return Response({"error": f"Missing text columns: {', '.join(missing_cols_text)}"}, status=status.HTTP_400_BAD_REQUEST)

        question_count = 0
        duplicate_questions = []

        try:
            with transaction.atomic():
                for index, row in df.iterrows():
                    question_text = str(get_column_data('Question_Text', row)) if get_column_data('Question_Text', row) is not None else None
                    question_image_data = None # Placeholder for image data if you're extracting from Excel
                    question_image_filename = None # Placeholder

                    options_text = [
                        str(get_column_data(f'Option{i}_Text', row)) if get_column_data(f'Option{i}_Text', row) is not None else None
                        for i in range(1, 5)
                    ]
                    options_image_data = [None] * 4 # Placeholders
                    options_image_filenames = [None] * 4 # Placeholders

                    correct_answer = str(get_column_data('Answer', row)).strip().lower().replace(" ", "")
                    subject_name = str(get_column_data('Subject', row))
                    category_name = str(get_column_data('Category', row))
                    difficulty_level_str = str(get_column_data('Difficulty', row))

                    subject, _ = Subject.objects.get_or_create(name=subject_name)
                    category, _ = Category.objects.get_or_create(name=category_name)

                    try:
                        difficulty_level = int(difficulty_level_str)
                        if not 1 <= difficulty_level <= 6:
                            return Response({"error": f"Invalid difficulty level '{difficulty_level_str}'. It must be between 1 and 6."}, status=status.HTTP_400_BAD_REQUEST)
                    except ValueError:
                        return Response({"error": f"Invalid difficulty level '{difficulty_level_str}'. It must be an integer."}, status=status.HTTP_400_BAD_REQUEST)

                    # Basic duplicate check based on text (you might need to adjust)
                    if Question.objects.filter(exam=exam, text=question_text).exists():
                        duplicate_questions.append(question_text if question_text else "Question with image")
                        continue

                    question = Question.objects.create(
                        exam=exam,
                        text=question_text,
                        image=None,
                        marks=1,
                        category=category,
                        difficulty_level=difficulty_level,
                        subject=subject
                    )
                    question_count += 1

                    # Placeholder for handling question image (if you implement image extraction)
                    # if question_image_data:
                    #     image_file = save_image_with_unique_name(question_image_data, "question_image.png", 'questions')
                    #     if image_file:
                    #         question.image.save(image_file.name.split('/')[-1], image_file)
                    #         question.save()

                    for i, option_text in enumerate(options_text):
                        is_correct = (f"option{i+1}".strip().lower().replace(" ", "") == correct_answer)
                        option = QuestionOption.objects.create(
                            question=question,
                            text=option_text,
                            image=None,
                            is_correct=is_correct
                        )
                        # Placeholder for handling option images
                        # if options_image_data[i]:
                        #     image_file = save_image_with_unique_name(options_image_data[i], f"option_{i+1}.png", 'options')
                        #     if image_file:
                        #         option.image.save(image_file.name.split('/')[-1], image_file)
                        #         option.save()

                exam.total_questions = Question.objects.filter(exam=exam).count()
                exam.save()

                if duplicate_questions:
                    return Response({"message": f"{question_count} questions created. Some were duplicates and skipped.", "duplicates": duplicate_questions}, status=status.HTTP_201_CREATED)
                else:
                    return Response({"message": f"{question_count} questions created successfully."}, status=status.HTTP_201_CREATED)

        except Exception as e:
            logging.error(f"An error occurred: {str(e)}")
            return Response({"error": f"An error occurred: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)



class TeacherListView(ListAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get_queryset(self):
        teachers = User.objects.filter(role='teacher').exclude(id=self.request.user.id)
        print("the teacher", teachers[0].username)
        # Filter users who are teachers (you might use a role or a specific flag)
        return teachers

class StudentListView(ListAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get_queryset(self):
        students = User.objects.filter(role='student').exclude(id=self.request.user.id)
        # print("the student", students[0].username)
        # Filter users who are teachers (you might use a role or a specific flag)
        return students

class QuestionHistoryByMonthView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    authentication_classes = [JWTAuthentication]

    def get(self, request, format=None):
        year = request.query_params.get('year')
        month = request.query_params.get('month')

        if not year or not month:
            return Response({'detail': 'Year and month are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            year = int(year)
            month = int(month)
            if month < 1 or month > 12:
                raise ValueError
        except (ValueError, TypeError):
            return Response({'detail': 'Invalid year or month.'}, status=status.HTTP_400_BAD_REQUEST)

        # Construct the start and end date for the month
        start_date = parse_date(f'{year}-{month:02d}-01')
        end_day = start_date.replace(day=28) + timedelta(days=4)  # Get the last day of the month
        end_date = end_day - timedelta(days=end_day.day)

        # Filter questions that are published and created within the date range
        questions = Question.objects.filter(
            created_at__range=(start_date, end_date),
            status='published'
        )

        # Serialize the results
        serializer = QuestionSerializer(questions, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    
class QuestionHistoryByTeacherMonthYearView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
    authentication_classes = [JWTAuthentication]

    def get(self, request, format=None):
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        teacher_id = request.query_params.get('teacher_id')

        # Validate parameters
        if not year or not month or not teacher_id:
            return Response({'detail': 'Year, month, and teacher_id are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            year = int(year)
            month = int(month)
            teacher_id = int(teacher_id)
            if month < 1 or month > 12:
                raise ValueError
        except (ValueError, TypeError):
            return Response({'detail': 'Invalid year, month, or teacher_id.'}, status=status.HTTP_400_BAD_REQUEST)
        
        
        # Verify teacher_id is a valid teacher
        if not User.objects.filter(id=teacher_id, role='teacher').exists():
            return Response({'detail': 'Invalid teacher_id.'}, status=status.HTTP_400_BAD_REQUEST)

        
        
        # Annotating month and year from created_at
        questions = Question.objects.annotate(
            question_month=ExtractMonth('created_at'),
            question_year=ExtractYear('created_at')
        ).filter(
            created_by_id=teacher_id
        )

        serializer = QuestionSerializer(questions, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class UserQuestionSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        month = request.data.get('month')
        year = request.data.get('year')
        user_id = request.data.get('user_id')
        print("hello world")
        print(month, year, user_id)
        # Validate parameters
        if not year or not month:
            return Response({'detail': 'Year and month are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            year = int(year)
            month = int(month)
            if month < 1 or month > 12:
                raise ValueError
        except ValueError:
            return Response({'detail': 'Invalid year or month.'}, status=status.HTTP_400_BAD_REQUEST)

        results = []

        if user_id:
            if isinstance(user_id, str) and user_id.lower() == "all":
                teachers = User.objects.filter(role='teacher')
                overall_summary = self.get_category_summary(year, month)
            else:
                try:
                    user_id = int(user_id)
                    teachers = User.objects.filter(id=user_id, role='teacher')
                except ValueError:
                    return Response({'detail': 'Invalid user ID.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            teachers = User.objects.filter(role='teacher')

        for teacher in teachers:
            total_questions = Question.objects.filter(
                Q(exam__created_by=teacher) & Q(exam__created_at__year=year, exam__created_at__month=month)
            ).count()

            category_counts = Question.objects.filter(
                Q(exam__created_by=teacher) & Q(exam__created_at__year=year, exam__created_at__month=month)
            ).values('category__name').annotate(
                question_count=Count('id')
            ).order_by('category__name')

            category_data = [
                {'category_name': item['category__name'], 'question_count': item['question_count']}
                for item in category_counts
            ]
            print(total_questions)

            results.append({
                'username': teacher.username,
                'total_questions': total_questions,
                'categories': category_data
            })

        if user_id and isinstance(user_id, str) and user_id.lower() == "all":
            # print(overall_summary)
            return Response({
                'overall_summary': overall_summary,
                'individual_teachers': results
            }, status=status.HTTP_200_OK)

        return Response(results, status=status.HTTP_200_OK)

    def get_category_summary(self, year, month):
        category_counts = Question.objects.filter(
            exam__created_at__year=year,
            exam__created_at__month=month
        ).values('category__name').annotate(
            question_count=Count('id')
        ).order_by('category__name')

        return [
            {'category_name': item['category__name'], 'question_count': item['question_count']}
            for item in category_counts
        ]

# Work Flow 
class ExamAttemptViewSet(viewsets.ViewSet):
    queryset = ExamAttempt.objects.all()
    serializer_class = ExamAttemptSerializer
    permission_classes = []
    
    
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def all_attempts(self, request):
        id = request.data.get('user_id', None)
        user = User.objects.get(id=id)
        print(id)
        time_period = request.query_params.get('time_period', 'all')

        # Query to get all the exams the user has attempted
        user_attempts = ExamAttempt.objects.filter(user=user).prefetch_related('exam__questions')

        # Filter by time period
        if time_period == 'weekly':
            user_attempts = user_attempts.filter(attempt_time__gte=timezone.now() - timedelta(weeks=1))
        elif time_period == 'monthly':
            user_attempts = user_attempts.filter(attempt_time__gte=timezone.now() - timedelta(days=30))
        elif time_period == 'yearly':
            user_attempts = user_attempts.filter(attempt_time__gte=timezone.now() - timedelta(days=365))

        if not user_attempts.exists():
            return Response({"message": "No exam attempts found for this user."}, status=status.HTTP_404_NOT_FOUND)

        # Prepare detailed data for each attempt
        attempts_data = []
        for attempt in user_attempts:
            exam = attempt.exam
            total_questions = exam.questions.count()
            correct_answers = attempt.total_correct_answers
            wrong_answers = total_questions - correct_answers  # Assuming no skipped questions
            
            # Calculate obtained marks and percentage
            total_marks = sum(question.marks for question in exam.questions.all())
            obtained_marks = correct_answers * (total_marks / total_questions)  # Assuming equal marks per question
            percentage = (obtained_marks / total_marks) * 100

            attempt_detail = {
                'exam_title': exam.title,
                'attempt_time': attempt.attempt_time,
                'total_questions': total_questions,
                'correct_answers': correct_answers,
                'wrong_answers': wrong_answers,
                'total_marks': total_marks,
                'obtained_marks': obtained_marks,
                'percentage': round(percentage, 2),
            }

            attempts_data.append(attempt_detail)

        return Response(attempts_data, status=status.HTTP_200_OK)
    
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def user_attempts(self, request):
        user = self.request.user
        print()
        exam_id = request.query_params.get('exam_id', None) # Get 'exam_id' from the query parameters
        user_id = request.query_params.get('user_id', None)
        # user = User.objects.get(id=user_id)
        # print(user_id, exam_id)
        if not exam_id:
            return Response({"error": "exam_id parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Validate that the exam exists
        try:
            exam = Exam.objects.get(exam_id=exam_id)
        except Exam.DoesNotExist:
            return Response({"error": "Exam not found."}, status=status.HTTP_404_NOT_FOUND)

        # Fetch all attempts by the user for the specific exam
        user_attempts = ExamAttempt.objects.filter(exam=exam, user=user)

        if not user_attempts.exists():
            return Response({"message": "No attempts found for this exam."}, status=status.HTTP_404_NOT_FOUND)

        # Serialize the attempts
        serializer = ExamAttemptSerializer(user_attempts, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)
    
    
    
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def highest_attempts(self, request):
        # user = self.request.user
        id = request.data.get('user_id', None)
        user = User.objects.get(id=id)

        # Query for exams the user has attempted
        attempted_exams = Exam.objects.filter(attempts__user=user).distinct()

        exams_data = []
        
        for exam in attempted_exams:
            # Get all attempts by the user for this exam
            # attempts = ExamAttempt.objects.filter(user=user, exam=exam)
            attempts = ExamAttempt.objects.filter(user=user, exam=exam).annotate(
                position=Window(
                    expression=Rank(),
                    order_by=F('total_correct_answers').desc()
                )
            )
            # Find the attempt with the highest correct answers
            highest_attempt = attempts.order_by('-total_correct_answers').first()
            unique_participants_count = ExamAttempt.objects.filter(exam=exam).values('user').distinct().count()
            
            # Add the highest attempt details to the response data
            exams_data.append({
                'exam_id': exam.exam_id,
                'exam_title': exam.title,
                'total_questions': exam.total_questions,
                'passed_marks':exam.pass_mark,
                'unique_participants': unique_participants_count,
                'highest_attempt': {
                    'attempt_id': highest_attempt.id,
                    'answered': highest_attempt.answered,
                    'total_correct_answers': highest_attempt.total_correct_answers,
                    'wrong_answers': highest_attempt.wrong_answers,
                    'passed': highest_attempt.passed,
                    'attempt_time': highest_attempt.attempt_time,
                    'position': highest_attempt.position
                }
            })

        return Response({'exams': exams_data}, status=status.HTTP_200_OK)

    

@login_required
@api_view(['GET'])
def user_attempts_by_month(request):
    user = request.user
    month = request.GET.get('month')
    year = request.GET.get('year')

    if not month or not year:
        return Response({"error": "Month and year parameters are required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        month = int(month)
        year = int(year)
    except ValueError:
        return Response({"error": "Invalid month or year."}, status=status.HTTP_400_BAD_REQUEST)

    # Validate month and year
    if month < 1 or month > 12 or year < 1:
        return Response({"error": "Month must be between 1 and 12 and year must be positive."}, status=status.HTTP_400_BAD_REQUEST)

    # Create timezone-aware dates
    start_date = timezone.make_aware(timezone.datetime(year, month, 1))
    end_date = (timezone.make_aware(timezone.datetime(year, month, 1)) + timezone.timedelta(days=31)).replace(day=1) - timezone.timedelta(seconds=1)

    # Filter attempts by the selected month and year
    attempts = ExamAttempt.objects.filter(user=user, timestamp__range=(start_date, end_date))

    if not attempts.exists():
        return Response({"message": "No attempts found for the selected month."})

    serializer = ExamAttemptSerializer(attempts, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@login_required
@api_view(['GET'])
@permission_classes([IsAdmin])
def user_exam_attempts_by_month(request):
    user_id = request.GET.get('user_id')
    month = request.GET.get('month')
    year = request.GET.get('year')

    if not user_id or not month or not year:
        return Response({"error": "User ID, month, and year parameters are required."})

    try:
        month = int(month)
        year = int(year)
        user = User.objects.get(id=user_id)
    except (ValueError, User.DoesNotExist):
        return Response({"error": "Invalid month, year, or user ID."})

    # Validate month and year
    if month < 1 or month > 12 or year < 1:
        return Response({"error": "Month must be between 1 and 12 and year must be positive."})

    # Create timezone-aware dates
    start_date = timezone.make_aware(timezone.datetime(year, month, 1))
    end_date = (timezone.make_aware(timezone.datetime(year, month, 1)) + timezone.timedelta(days=31)).replace(day=1) - timezone.timedelta(seconds=1)

    # Filter attempts by the selected month and year
    attempts = ExamAttempt.objects.filter(user=user, timestamp__range=(start_date, end_date))

    if not attempts.exists():
        return Response({"message": "No attempts found for the selected month for this user."})

    serializer = ExamAttemptSerializer(attempts, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)



class ExamSubjectsQuestionCountView(APIView):
    def get(self, request, exam_id):
        # Retrieve the specific exam (this step is optional if you only need questions)
        exam = get_object_or_404(Exam, exam_id=exam_id)
        
        # Get subjects and question counts for the specified exam
        subjects_with_question_count = (
            Question.objects.filter(exams=exam)
            .values('subject__name')  # Group by subject name
            .annotate(question_count=Count('id'))  # Count questions per subject
        )

        # Rename `subject__name` to `subject_name` for better readability
        results = [
            {'subject_name': item['subject__name'], 'question_count': item['question_count']}
            for item in subjects_with_question_count
        ]

        # Serialize the results
        serializer = SubjectQuestionCountSerializer(results, many=True)
        
        return Response(serializer.data)
    
    
    
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exam_leaderboard_view(request, exam_id):
    # Retrieve the specific exam, or return a 404 if not found
    try:
        exam = Exam.objects.get(pk=exam_id)
    except Exam.DoesNotExist:
        return Response({"detail": "Exam not found."}, status=404)
    
    # Query leaderboard with cumulative questions, score, and correct answers
    leaderboard = (
        Leaderboard.objects
        .filter(exam=exam)
        .values('user')
        .annotate(
            username=F('user__username'),
            cumulative_questions=Sum('total_questions'),
            cumulative_score=Sum('score'),
            total_correct=ExamAttempt.objects.filter(user=F('user'), exam=exam).aggregate(total_correct=Sum('total_correct_answers'))['total_correct'] or 0
        )
        .order_by('-cumulative_score')[:10]  # Top 10 users by score
    )
    
    # Serialize data
    serializer = ResultSerializer(leaderboard, many=True)
    return Response(serializer.data)








#User summury data


class UserExamSummaryAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, id):
        # userid = request.body.get('user_id')
        print(id)
        user = User.objects.get(id=id)
        
        print(user)
        # Get all attempts by the user
        attempts = ExamAttempt.objects.filter(user=user)
       
        # Aggregated data
        total_attempts = attempts.count()
        total_passed = attempts.filter(passed=True).count()
        total_failed = total_attempts - total_passed

        # Calculate total questions by multiplying the exam's total questions by the number of times the user attempted each exam
        unique_exams = set(attempt.exam for attempt in attempts)
        total_questions = sum(exam.total_questions * attempts.filter(exam=exam).count() for exam in unique_exams)
               
        total_answered = attempts.aggregate(total=Sum('answered'))['total'] or 0
        total_correct_answers = attempts.aggregate(total=Sum('total_correct_answers'))['total'] or 0
        total_wrong_answers = attempts.aggregate(total=Sum('wrong_answers'))['total'] or 0

        # Calculate unanswered questions
        total_unanswered = total_questions - total_answered

        # Structure the data for response
        data = {
            "username": user.username,
            "total_attempts": total_attempts,
            "total_passed": total_passed,
            "total_failed": total_failed,
            "total_questions": total_questions,
            "total_answered": total_answered,
            "total_correct_answers": total_correct_answers,
            "total_wrong_answers": total_wrong_answers,
            "total_unanswered": total_unanswered,
        }

        return Response(data)

class UserAnswerViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'], url_path='all-submitted-questions')
    def all_submitted_questions(self, request):
        id = request.data.get('user_id')
        user = User.objects.get(id=id)

        # Fetch all UserAnswer instances for the user
        answered_questions = UserAnswer.objects.filter(exam_attempt__user=user)

        # Filter for correct and incorrect answers
        correct_answers = answered_questions.filter(is_correct=True)
        wrong_answers = answered_questions.filter(is_correct=False)

        # Serialize the data
        correct_answers_data = UserAnswerSerializer(correct_answers, many=True, context={'include_options': True}).data
        wrong_answers_data = UserAnswerSerializer(wrong_answers, many=True, context={'include_options': True}).data
        submitted_questions_data = UserAnswerSerializer(answered_questions, many=True, context={'include_options': True}).data
        return Response({
            "submitted_questions": submitted_questions_data,
            "correct_answers": correct_answers_data,
            "wrong_answers": wrong_answers_data
        }, status=status.HTTP_200_OK)
        
        
        
        
        
        
        
# Exam Create

class ExamCreateView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def post(self, request, *args, **kwargs):
        exam_title = request.data.get('exam_title')
        total_questions = int(request.data.get('total_questions', 10))  # Default 10 questions
        total_marks = int(request.data.get('total_marks', 100))  # Default 100 marks
        pass_mark = int(request.data.get('pass_mark', 50))  # Default 50% pass mark
        last_date = request.data.get('last_date', None)
        duration = int(request.data.get('duration', 60))  # Duration in minutes
        negative_marks = request.data.get('negative_marks', 0)
        starting_time = request.data.get('starting_time', None)
        exam_type = request.data.get('exam_type', 'question_bank')  # Default type is 'question_bank'
        difficulty_levels = request.data.get('difficulty_levels', {})
        # print("difficulty: ", type(difficulty_levels))
        
        category_name = request.data.get('category')
       
        try:
            difficulty_levels = json.loads(difficulty_levels)
            print(difficulty_levels)
        except (TypeError, json.JSONDecodeError):
            return Response({"error": "Difficulty levels are invalid."}, status=status.HTTP_400_BAD_REQUEST)
        
        # Only validate difficulty percentages if exam type is 'question_bank'
        if exam_type == 'question_bank':
            # Ensure percentages add up to 100
            total_percentage = sum(difficulty_levels.values())
            if total_percentage != 100:
                return Response({"error": f"The total percentage of difficulty levels must equal 100, but got {total_percentage}."}, status=status.HTTP_400_BAD_REQUEST)

        
        # Basic validations for required fields
        if not exam_title:
            return Response({"error": "Exam title is required."}, status=status.HTTP_400_BAD_REQUEST)
        if not exam_type:
            return Response({"error": "Exam type is required (either 'file' or 'question_bank')."}, status=status.HTTP_400_BAD_REQUEST)
        
        category = None
        if category_name:
            exam_category, _ = ExamCategory.objects.get_or_create(name=category_name)
        
        
        selected_questions = []

        # Process exam questions based on exam type
        if exam_type == 'file':
            selected_questions = self._process_file_upload(request)
        elif exam_type == 'question_bank':
            print(type(difficulty_levels))
            selected_questions = self._fetch_questions_from_bank(total_questions, difficulty_levels)
        else:
            return Response({"error": "Invalid exam type. Choose either 'file' or 'question_bank'."}, status=status.HTTP_400_BAD_REQUEST)

        if isinstance(selected_questions, Response):  # Handle error response from helper functions
            return selected_questions

        duration_minutes = timedelta(minutes=duration)
        # Step 2: Create the Exam
        exam = Exam.objects.create(
            title=exam_title,
            total_questions=len(selected_questions),
            total_mark=len(selected_questions),
            pass_mark=pass_mark,
            last_date=last_date,
            duration=duration_minutes,  # Convert minutes to seconds
            negative_mark=negative_marks,
            starting_time=parse_datetime(starting_time) if starting_time else None,
            created_by=request.user,
            category=exam_category
        )

        
        
        # Step 3: Associate Questions with the Exam
        exam.questions.set(selected_questions)

       
        print(difficulty_levels)

        # Step 5
        difficulty_percentages = {
            'difficulty1_percentage': difficulty_levels.get('1', 0),
            'difficulty2_percentage': difficulty_levels.get('2', 0),
            'difficulty3_percentage': difficulty_levels.get('3', 0),
            'difficulty4_percentage': difficulty_levels.get('4', 0),
            'difficulty5_percentage': difficulty_levels.get('5', 0),
            'difficulty6_percentage': difficulty_levels.get('6', 0)
        }

        # Create ExamDifficulty object
        ExamDifficulty.objects.create(
            exam=exam,
            **difficulty_percentages
        )
        
        # Step 5: Set Exam Status and Initialize Leaderboard
        status_label = 'student' if request.user.role == 'student' else 'draft'
        Status.objects.create(exam=exam, status=status_label, user=request.user)

        # Leaderboard.objects.create(exam=exam, score=0)

        return Response({"message": f"Exam '{exam.title}' created successfully with {len(selected_questions)} questions."}, status=status.HTTP_201_CREATED)

    def _process_file_upload(self, request):
        print(request.FILES)
        """Handles the 'file' type exam creation by processing an uploaded Excel file."""
        if 'file' not in request.FILES:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return Response({"error": f"Error reading file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        required_columns = ['Question', 'Option1', 'Option2', 'Option3', 'Option4', 'Answer', 'Options_num', 'Category', 'Difficulty', 'Subject']
        missing_cols = [col for col in required_columns if col.lower() not in [col.lower() for col in df.columns]]
        if missing_cols:
            return Response({"error": f"Missing columns: {', '.join(missing_cols)}"}, status=status.HTTP_400_BAD_REQUEST)

        selected_questions = []
        for _, row in df.iterrows():
            row = {col.lower(): value for col, value in row.items()}
            try:
                question = self._create_question_from_row(row)
                selected_questions.append(question)
            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return selected_questions

    def _create_question_from_row(self, row):
        """Creates a question and its options from a given row in the DataFrame."""
        question_text = row['question']
        options = [row['option1'], row['option2'], row['option3'], row['option4']]
        correct_answer = row['answer'].strip().lower().replace(" ", "")
        
        subject, _ = Subject.objects.get_or_create(name=row['subject'])
        category, _ = Category.objects.get_or_create(name=row['category'])
        try:
            difficulty_level = int(row['difficulty'])
           
            if difficulty_level not in range(1, 7):
                raise ValueError(f"Invalid difficulty level {difficulty_level}. It must be between 1 and 6.")
        except ValueError:
            raise ValueError(f"Invalid difficulty level '{row['difficulty']}'. It must be an integer between 1 and 6.")

        question, created = Question.objects.get_or_create(
            text=question_text,
            category=category,
            subject=subject,
            defaults={'marks': 1, 'difficulty_level': difficulty_level}
        )

        if created:
            for i, option_text in enumerate(options, start=1):
                is_correct = (f"option{i}".strip().lower() == correct_answer)
                QuestionOption.objects.create(question=question, text=option_text, is_correct=is_correct)
        else:
            print(f"Question '{question_text}' already exists in the database.")

        return question

    def _fetch_questions_from_bank(self, num_questions, difficulty_levels):
        """Fetches a set number of questions from the question bank based on difficulty levels."""
        
        try:
            with transaction.atomic():
                # difficulty_levels = json.loads(difficulty_levels)

                # Ensure percentages add up to 100
                if sum(difficulty_levels.values()) != 100:
                    return Response({"error": "Difficulty percentages must add up to 100%."}, status=status.HTTP_400_BAD_REQUEST)

                selected_questions = []
               
                for level, percentage in difficulty_levels.items():
                    # Calculate the number of questions to fetch for this level
                    num_level_questions = round(num_questions * (percentage / 100))

                    # Filter questions by difficulty level and randomly order them
                    level_questions = Question.objects.filter(difficulty_level=level).order_by('?')[:num_level_questions]

                    # Check if we have enough questions at this difficulty level
                    if level_questions.count() < num_level_questions:
                        return Response(
                            {"error": f"Not enough questions available for difficulty level '{level}' to meet the requested number."},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                    selected_questions.extend(level_questions)

                # Shuffle selected questions for random order
                random.shuffle(selected_questions)
                
                # Ensure the total number of selected questions does not exceed `num_questions`
                return selected_questions[:num_questions]

        except Exception as e:
            return Response({"error": f"An error occurred while fetching questions: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        
        
        
        
class OrganizationViewSet(viewsets.ModelViewSet):
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer
    
class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    
class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    






from rest_framework import viewsets, status
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile
from django.db import transaction
from io import BytesIO
from PIL import Image






class PastExamViewSet(viewsets.ModelViewSet):
    queryset = PastExam.objects.all().order_by("-exam_date")
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return PastExam.objects.all()
        # user = self.request.user
        # return PastExam.objects.filter(created_by=user).order_by("-exam_date")
    
    def get_serializer_class(self):
        if self.action == "list":
            return PastExamListSerializer
        elif self.action in ["create", "update", "partial_update"]:
            return PastExamCreateSerializer
        return PastExamSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data, context={"request": request})

        if serializer.is_valid():
            past_exam = serializer.save()
            past_exam.refresh_from_db()

            if "file" in request.FILES:
                file = request.FILES["file"]
                result = self.process_questions(file, past_exam)

                if result and "error" in result:
                    return Response(result, status=status.HTTP_400_BAD_REQUEST)
                elif result:
                    return Response(result, status=status.HTTP_201_CREATED)
                else:
                    return Response({"error": "Unknown error occurred during question processing."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            return Response(PastExamSerializer(past_exam).data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    
  
    
    def process_questions(self, file, past_exam):
        # 1. Read Excel file with pandas and openpyxl for image extraction
        try:
            logging.info(f"Starting to process the file for PastExam ID: {past_exam.id}")

            # Create a temporary file on disk and write the uploaded file content to it
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file.write(file.read())
                temp_file_path = temp_file.name

            # --- Load with openpyxl First for Image Mapping ---
            try:
                workbook = openpyxl.load_workbook(temp_file_path)
                sheet = workbook.active
                image_map = {}
                for image in sheet._images:
                    if hasattr(image, "anchor") and hasattr(image.anchor, '_from'):
                         # Openpyxl uses 0-based indexing for row/col internally
                        cell = image.anchor._from
                        # Cell reference needs 1-based indexing
                        cell_ref = f"{get_column_letter(cell.col + 1)}{cell.row + 1}"
                        image_map[cell_ref] = image
                    else:
                        logging.warning("Found image without a valid anchor.")
                logging.info(f"Found {len(image_map)} images mapped to cells.")
            except Exception as e:
                 logging.error(f"Error loading workbook with openpyxl or mapping images: {e}")
                 # Clean up temp file before returning
                 try:
                     os.remove(temp_file_path)
                 except OSError:
                     pass # Ignore error if file couldn't be removed
                 return {"error": f"Error processing Excel file structure or images: {e}"}


            # --- Load with pandas for Data Iteration ---
            # Use string dtype to prevent pandas from guessing types (like numbers)
            df = pd.read_excel(temp_file_path, dtype=str)
            # Explicitly fill NaN values with empty strings AFTER loading
            df = df.fillna("")


            # --- Clean up temp file ---
            try:
                os.remove(temp_file_path)
                logging.info(f"Temporary file {temp_file_path} removed.")
            except OSError as e:
                logging.warning(f"Could not remove temporary file {temp_file_path}: {e}")


            # # Ensure all data is string - Redundant if loaded with dtype=str
            # for column in df.columns:
            #     df[column] = df[column].astype(str).str.strip() # Also strip whitespace

            # 2. Convert all column names to lowercase for consistency
            df.columns = [str(col).lower().strip() for col in df.columns]

            # 3. Determine optional fields & Validate required columns
            # Check for 'question' column existence more robustly
            question_column_name = next((col for col in df.columns if 'question' in col), None)
            if not question_column_name:
                 return {"error": "Missing critical column: 'question' column not found."}

            required_columns = [question_column_name] # Start with the actual question column found
            has_answer = "answer" in df.columns
            has_difficulty = "difficulty" in df.columns
            has_category = "category" in df.columns
            
            if has_answer:
                required_columns.append("answer")
            # Add similar checks for difficulty/category if they are mandatory

            # Validate columns needed for processing this row
            # This validation might be too strict depending on requirements
            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
               return {"error": f"Missing columns: {', '.join(missing_cols)}"}


            # 5. Option style support
            standard_options = ["option1", "option2", "option3", "option4"]
            bengali_options = ["ক", "খ", "গ", "ঘ"]
            # Detect available option columns dynamically
            available_std_options = [opt for opt in standard_options if opt in df.columns]
            available_beng_options = [opt for opt in bengali_options if opt in df.columns]

            option_format_detected = False
            if len(available_std_options) >= 1 : # Check if at least 2 standard options exist
                 detected_option_cols = available_std_options
                 answer_map = {"ক": "option1", "খ": "option2", "গ": "option3", "ঘ": "option4"}
                 # Add direct mapping for standard options if answer column uses them
                 answer_map.update({opt: opt for opt in standard_options})
                 option_format_detected = True
                 logging.info("Detected standard option format (option1, option2...).")
            elif len(available_beng_options) >= 1: # Check if at least 2 bengali options exist
                 detected_option_cols = available_beng_options
                 answer_map = {opt: opt for opt in bengali_options}
                 option_format_detected = True
                 logging.info("Detected Bengali option format (ক, খ...).")
            else:
                 # No recognizable options found or insufficient options
                 logging.warning("Could not detect sufficient option columns (need at least 2 of standard or Bengali format).")
                 # Decide if this is an error or if questions without options are allowed
                 # return {"error": "Insufficient option columns found in the file."}
                 detected_option_cols = [] # Proceed without options if allowed
                 answer_map = {}


            # 6. Initialize counters
            current_subject = None
            question_count = 0
            updated_questions = 0
            skipped_rows = 0


            # 7. Atomic transaction for rollback on error
            with transaction.atomic():
                last_valid_subject = None # Keep track of the last subject found
                for index, row in df.iterrows():
                    # row = row.fillna("") # Already done after loading df

                    excel_row_num = index + 2 # For user-friendly messages

                    # 7.1 Detect subject - Allow subject to persist from previous rows
                    subject_text = str(row.get("subject", "")).strip()
                    if subject_text:
                        current_subject, _ = Subject.objects.get_or_create(name=subject_text)
                        last_valid_subject = current_subject # Update the last valid subject
                    elif last_valid_subject:
                         current_subject = last_valid_subject # Use the last known subject
                    else:
                        # No subject found yet and none in this row, cannot proceed
                        logging.error(f"Skipping row {excel_row_num}: Subject missing and no previous subject found.")
                        skipped_rows += 1
                        continue # Skip this row


                    # 7.2 Detect question content (Text and Image)
                    question_cell = f"{get_column_letter(df.columns.get_loc(question_column_name) + 1)}{excel_row_num}"
                    # Get potential text, strip whitespace. Will be "" if empty or only whitespace.
                    original_question_text = str(row.get(question_column_name, "")).strip()
                    question_image_data = self.get_image_data_from_map(image_map.get(question_cell))
                    # if original_question_text:
                    #     print(original_question_text)
                    # --- *** NEW CHECK: Skip if Question Text AND Image are missing *** ---
                    if not original_question_text and not question_image_data:
                        logging.warning(f"Skipping row {excel_row_num}: Question text and image are both missing.")
                        skipped_rows += 1
                        continue # Move to the next row
                    # --- *** END NEW CHECK *** ---

                    # Determine the text to store in the DB (None if image exists)
                    question_text_for_db = None if question_image_data else original_question_text


                    # 7.3 Detect category
                    category = None
                    category_text = str(row.get("category", "")).strip()
                    if has_category and category_text:
                        category, _ = Category.objects.get_or_create(name=category_text)


                    # 7.4 Detect difficulty level
                    difficulty_level = None
                    difficulty_text = str(row.get("difficulty", "")).strip()
                    if has_difficulty and difficulty_text:
                        try:
                            difficulty_level = int(difficulty_text)
                        except ValueError:
                            logging.error(f"Invalid difficulty level '{difficulty_text}' at row {excel_row_num}. Setting to None.")
                            # Decide if this should be a hard error:
                            # transaction.set_rollback(True)
                            # return {"error": f"Invalid difficulty level '{difficulty_text}' at row {excel_row_num}"}
                            difficulty_level = None # Or skip row, or default value


                    # 7.5 Create or update question
                    # Use question_text_for_db which is None if image exists
                    if question_text_for_db:
                        
                         question, created = Question.objects.get_or_create(
                             text=question_text_for_db,
                            #  subject=current_subject, # Match subject too? Decide based on desired uniqueness.
                             defaults={
                                 
                                 "marks": 1, # Assuming default marks
                                 "category": category,
                                 "difficulty_level": difficulty_level
                             }
                         )
                         if created:
                             question_count += 1
                             question.subject = current_subject
                         else:
                             # Update existing question if necessary
                             updated = False
                             if not question.category and category:
                                 question.category = category
                                 updated = True
                             if question.difficulty_level is None and difficulty_level is not None:
                                 question.difficulty_level = difficulty_level
                                 updated = True
                             # Add other fields to update if needed
                             if updated:
                                 question.save()
                                 updated_questions += 1

                    else: # No text (must be an image question) -> Always create new
                         question = Question.objects.create(
                             text=None,
                             subject=current_subject,
                             marks=1,
                             category=category,
                             difficulty_level=difficulty_level
                             # Image is saved later
                         )
                         question_count += 1

                    existing_usage = QuestionUsage.objects.filter(
                        question=question,
                        past_exam=past_exam,
                        year=past_exam.exam_date.year
                    ).first()

                    if not existing_usage:
                        QuestionUsage.objects.create(
                            question=question,
                            past_exam=past_exam,
                            year=past_exam.exam_date.year
                        )
                    else:
                        logging.info(f"QuestionUsage already exists for question ID {question.id}, past exam ID {past_exam.id}, and year {past_exam.exam_date.year} at row {excel_row_num}. Skipping creation.")
                    # 7.6 Link question to PastExam
                    past_exam_question = PastExamQuestion.objects.create(
                        exam=past_exam,
                        question=question,
                        order=index + 1 # Use DataFrame index for order
                    )


                    # 7.7 Save question image if exists
                    if question_image_data:
                        filename = f"question_image_{question.id}_{uuid.uuid4().hex[:8]}.png"
                        image_file = self.save_image_to_field(question_image_data, filename)
                        if image_file:
                            question.image.save(image_file.name, image_file, save=True)
                            # question.save() # save=True above handles this
                        else:
                             logging.error(f"Failed to process or save image for question at row {excel_row_num}")
                             # Decide if this is a fatal error for the row/transaction


                    # 7.8 Option format was determined earlier (detected_option_cols, answer_map)


                    # 7.9 Create options if format was detected
                    options_data_for_answer_check = [] # Store (option_key, option_obj)
                    
                    if detected_option_cols:
                        for i, option_key in enumerate(detected_option_cols):
                            option_text_original = str(row.get(option_key, "")).strip()
                            col_letter = get_column_letter(df.columns.get_loc(option_key) + 1)
                            option_cell = f"{col_letter}{excel_row_num}"
                            option_image_data = self.get_image_data_from_map(image_map.get(option_cell))

                            # If image exists, use None for text in DB logic
                            option_text_for_db = None if option_image_data else option_text_original

                            # Skip creating option if both text and image are missing for the option itself
                            if not option_text_original and not option_image_data:
                                logging.warning(f"Skipping empty option '{option_key}' for question at row {excel_row_num}.")
                                continue # Skip this specific option

                            # Conditional Creation Logic for Option
                            if option_text_for_db:
                                # Text exists: Try to reuse within the same question
                                option_obj, created = QuestionOption.objects.get_or_create(
                                    question=question,
                                    text=option_text_for_db
                                )
                            else:
                                # No text (image-only or blank): Always create new for this question
                                option_obj = QuestionOption.objects.create(
                                    question=question,
                                    text=None
                                )

                            # Link option to the specific instance in this exam
                            PastExamQuestionOption.objects.create(
                                question=past_exam_question,
                                option=option_obj
                            )

                            # Save option image if it exists
                            if option_image_data:
                                opt_filename = f"option_image_{option_obj.id}_{uuid.uuid4().hex[:8]}.png"
                                opt_image_file = self.save_image_to_field(option_image_data, opt_filename)
                                if opt_image_file:
                                     # Ensure we save the image to the correct object (newly created or retrieved)
                                    option_obj.image.save(opt_image_file.name, opt_image_file, save=True)
                                else:
                                     logging.error(f"Failed to process or save image for option '{option_key}' at row {excel_row_num}")


                            # Store for answer checking - use the original key from the loop
                            options_data_for_answer_check.append((option_key, option_obj))


                    # 7.10 Mark correct answer
                    if has_answer and detected_option_cols: # Only if answer column and options exist
                        correct_answer_raw = str(row.get("answer", "")).strip()
                        # Normalize answer (lowercase, remove spaces) only if it's not empty
                        correct_answer_normalized = correct_answer_raw.lower().replace(" ", "") if correct_answer_raw else None

                        if correct_answer_normalized and correct_answer_normalized in answer_map:
                             correct_key = answer_map.get(correct_answer_normalized) # Find the corresponding option column key
                             found_correct = False
                             for opt_key, opt_obj in options_data_for_answer_check:
                                 if opt_key == correct_key:
                                     opt_obj.is_correct = True
                                     opt_obj.save()
                                     found_correct = True
                                     break # Found the correct one
                             if not found_correct:
                                  logging.warning(f"Correct answer '{correct_answer_raw}' provided at row {excel_row_num}, but corresponding option '{correct_key}' was not found or processed.")
                        elif correct_answer_raw: # Answer provided but not parsable/mappable
                             logging.warning(f"Could not map correct answer '{correct_answer_raw}' at row {excel_row_num} to a valid option key.")
                        # else: No answer provided, which might be acceptable


                # Final save for PastExam (if any fields were updated, otherwise optional)
                past_exam.save()

            # 8. Return summary
            return {
                "message": f"Processing complete. {question_count} questions added, {updated_questions} updated, {skipped_rows} rows skipped.",
                "past_exam_id": past_exam.id
            }
        except Exception as e:
            # Log the full traceback for detailed debugging
            logging.exception(f"Critical error during question processing: {str(e)}") # Use logging.exception to include traceback
            # Ensure cleanup if temp file path was generated
            if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except OSError:
                    pass
            return {"error": f"An unexpected error occurred: {str(e)}"}



    def get_image_data_from_map(self, image):
        if not image:
            return None
        try:
            img_byte_arr = BytesIO()
            # Ensure the image object has the _data method
            if hasattr(image, '_data'):
                data = image._data()
                img_byte_arr.write(data)
                return img_byte_arr.getvalue()
            else:
                print(f"Error extracting image from cell: Unknown -> Image object has no '_data' attribute")
                return None
        except Exception as e:
            print(f"Error extracting image from cell: {getattr(image.anchor, 'from', 'Unknown')} -> {e}")
            return None

    # 💾 Helper: Convert raw image data to Django ContentFile
    
    
    # def save_image_to_field(self, image_data, filename):
    #     try:
    #         img = Image.open(BytesIO(image_data))
    #         img_format = img.format or "PNG"
    #         img_io = BytesIO()
    #         img.save(img_io, format=img_format)
    #         return ContentFile(img_io.getvalue(), name=filename)
    #     except Exception as e:
    #         print(f"Error saving image: {e}")
    #         return None
    
        
    def save_image_to_field(self, image_data, filename):
        try:
            img = Image.open(BytesIO(image_data))
            # Ensure image is converted to RGB if it's P mode with transparency
            if img.mode == 'P' and 'transparency' in img.info:
                 img = img.convert('RGBA')
            elif img.mode == 'RGBA':
                 # If saving as JPEG, need to convert RGBA to RGB
                 # For PNG, RGBA is fine. Let's assume PNG for simplicity or check format.
                 pass # Keep as RGBA for PNG
            elif img.mode != 'RGB':
                 img = img.convert('RGB')

            img_format = img.format or "PNG" # Default to PNG
            # Ensure filename has the correct extension if needed, though ContentFile doesn't strictly require it
            # filename = f"{filename.split('.')[0]}.{img_format.lower()}"

            img_io = BytesIO()
            img.save(img_io, format=img_format)
            img_io.seek(0) # Reset buffer position after writing
            return ContentFile(img_io.getvalue(), name=filename)
        except Exception as e:
            logging.error(f"Error processing or saving image '{filename}': {e}")
            return None

    # 💾 Helper: Convert raw image data to Django ContentFile
    # def save_image_to_field(self, image_data, filename):
    #     try:
    #         img = Image.open(BytesIO(image_data))
    #         img_format = img.format or "PNG"
    #         img_io = BytesIO()
    #         img.save(img_io, format=img_format)
    #         return ContentFile(img_io.getvalue(), name=filename)
    #     except Exception as e:
    #         print(f"Error saving image: {e}")
    #         return None


    


    @action(detail=False, methods=["get"])
    def list_by_organization(self, request):
        """List past exams filtered by organization ID."""
        organization_id = request.query_params.get("organization_id")
        if not organization_id:
            return Response({"error": "organization_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        exams = PastExam.objects.filter(organization_id=organization_id)
        serializer = self.get_serializer(exams, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def questions(self, request, pk=None):
        """Retrieve questions for a specific past exam."""
        past_exam = get_object_or_404(PastExam, pk=pk)
        questions = past_exam.questions.all()
        return Response({"questions": [q.text for q in questions]})


class UserPastExamListAPIView(generics.ListAPIView):
    serializer_class = PastExamSerializer
    permission_classes = [IsAuthenticated]  # Optional: restrict to logged-in users

    def get_queryset(self):
        user = self.request.user
        
        exams = PastExam.objects.filter(
            created_by=user
        ).order_by('-exam_date')
        print(exams)
        return exams

class PastExamListView(APIView):
    def get(self, request):
        exam_id = request.GET.get('exam_id')  # Fetch from query params

        if exam_id:
            exam = get_object_or_404(PastExam, id=exam_id)
            serializer = PastExamListSerializer(exam)
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        organization = request.GET.get('organization')
        department = request.GET.get('department')
        position = request.GET.get('position')
        exam_date = request.GET.get('exam_date')

        # exams = PastExam.objects.filter(is_published=True)
        exams = PastExam.objects.all()
        if organization:
            exams = exams.filter(organization__id=organization)
        if department:
            exams = exams.filter(department__id=department)
        if position:
            exams = exams.filter(position__id=position)
        if exam_date:
            exams = exams.filter(exam_date=exam_date)

        if not (organization or department or position or exam_date):
            exams = list(exams)
            random.shuffle(exams)

        serializer = PastExamListSerializer(exams, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class PastExamDetailView(generics.RetrieveAPIView):
    """
    API view to get details of a single past exam.
    """
    queryset = PastExam.objects.all()
    serializer_class = PastExamSerializer
    permission_classes = [IsAuthenticated]
    
    
class SubmitPastExamAttemptView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, exam_id):
        try:
            user = request.user
            past_exam = get_object_or_404(PastExam, id=exam_id)
            print(past_exam)
            submitted_answers = request.data.get("answers", [])
            print("Submitted answers:", submitted_answers)

            if not submitted_answers:
                return Response({"message": "No answers submitted."}, status=status.HTTP_400_BAD_REQUEST)

            correct_count = 0
            wrong_count = 0
            answered_count = 0

            with transaction.atomic():
                # Create a new exam attempt
                attempt = PastExamAttempt.objects.create(
                    user=user,
                    past_exam=past_exam,
                    total_questions=past_exam.questions.count()
                )

                for answer in submitted_answers:
                    question_id = answer.get("question_id")
                    selected_option_id = answer.get("selected_option_id")
                    print(f"Question ID: {question_id}, Selected Option ID: {selected_option_id}")

                    # Skip if question or answer is missing, or if the selected_option_id is invalid
                    if not question_id or not selected_option_id or selected_option_id in [None, 'none', '']:
                        continue

                    # Ensure selected_option_id is a valid integer
                    try:
                        print("hello", selected_option_id)
                        selected_option_id = int(selected_option_id)
                    except (ValueError, TypeError):
                        print(f"Invalid selected_option_id: {selected_option_id}")
                        continue

                    # Ensure the option exists in the database
                    try:
                        option = QuestionOption.objects.get(id=selected_option_id)
                    except QuestionOption.DoesNotExist:
                        print(f"Option with ID {selected_option_id} does not exist.")
                        continue

                    # Increment the answer counters
                    answered_count += 1
                    if option.is_correct:
                        correct_count += 1
                    else:
                        wrong_count += 1

                # Check if the user passed the exam
                is_passed = correct_count >= past_exam.pass_mark

                # Save the exam attempt results
                attempt.answered_questions = answered_count
                attempt.correct_answers = correct_count
                attempt.wrong_answers = wrong_count
                attempt.calculate_score()
                attempt.save()

            # Return response with the exam results
            return Response({
                "message": "Exam submitted successfully!",
                "total_questions": past_exam.questions.count(),
                "answered_questions": answered_count,
                "correct_answers": correct_count,
                "wrong_answers": wrong_count,
                "is_passed": is_passed,
                "score": attempt.score
            }, status=status.HTTP_200_OK)

        except Exception as e:
            print("Unexpected error:", e)
            traceback.print_exc()
            return Response({
                "error": "Something went wrong.",
                "detail": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class UpdateQuestionExplanationView(APIView):
    parser_classes = (MultiPartParser, FormParser)
    permission_classes = [IsAuthenticated]  # Only admin can upload

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({"error": "No file provided"}, status=400)

        try:
            # Read the Excel file
            df = pd.read_excel(file, engine='openpyxl')

            # Ensure required columns exist
            required_columns = {"text", "explanation"}
            if not required_columns.issubset(df.columns):
                return JsonResponse({"error": f"Missing required columns: {required_columns - set(df.columns)}"}, status=400)

            updated_questions = 0

            for _, row in df.iterrows():
                question_text = str(row["text"]).strip()
                explanation = str(row["explanation"]).strip() if pd.notna(row["explanation"]) else None

                try:
                    question = Question.objects.get(text=question_text)
                    if not question.explanation and explanation:
                        question.explanation = explanation
                        question.save()
                        updated_questions += 1

                except ObjectDoesNotExist:
                    # If question doesn't exist, skip it (or create a new one if needed)
                    pass

            return JsonResponse({"message": f"{updated_questions} questions updated."})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)



# past Exam
class PastExamAttemptViewSet(viewsets.ViewSet):
    queryset = PastExamAttempt.objects.all()
    serializer_class = PastExamSerializer
    permission_classes = []

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def all_attempts(self, request):
        user_id = request.data.get('user_id', None)
        user = User.objects.get(id=user_id)
        time_period = request.query_params.get('time_period', 'all')

        user_attempts = PastExamAttempt.objects.filter(user=user).prefetch_related('past_exam')

        if time_period == 'weekly':
            user_attempts = user_attempts.filter(attempt_time__gte=timezone.now() - timedelta(weeks=1))
        elif time_period == 'monthly':
            user_attempts = user_attempts.filter(attempt_time__gte=timezone.now() - timedelta(days=30))
        elif time_period == 'yearly':
            user_attempts = user_attempts.filter(attempt_time__gte=timezone.now() - timedelta(days=365))

        if not user_attempts.exists():
            return Response({"message": "No past exam attempts found for this user."}, status=status.HTTP_404_NOT_FOUND)

        attempts_data = []
        for attempt in user_attempts:
            past_exam = attempt.past_exam
            percentage = (attempt.correct_answers / attempt.total_questions) * 100 if attempt.total_questions > 0 else 0

            attempt_detail = {
                'exam_title': past_exam.title,
                'attempt_time': attempt.attempt_time,
                'total_questions': attempt.total_questions,
                'answered_questions': attempt.answered_questions,
                'correct_answers': attempt.correct_answers,
                'wrong_answers': attempt.wrong_answers,
                'score': attempt.score,
                'percentage': round(percentage, 2),
            }
            attempts_data.append(attempt_detail)

        return Response(attempts_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def user_attempts(self, request):
        user = request.user
        past_exam_id = request.query_params.get('past_exam_id', None)

        if not past_exam_id:
            return Response({"error": "past_exam_id parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            past_exam = PastExam.objects.get(id=past_exam_id)
        except PastExam.DoesNotExist:
            return Response({"error": "Past exam not found."}, status=status.HTTP_404_NOT_FOUND)

        user_attempts = PastExamAttempt.objects.filter(user=user, past_exam=past_exam)

        if not user_attempts.exists():
            return Response({"message": "No attempts found for this past exam."}, status=status.HTTP_404_NOT_FOUND)

        serializer = PastExamAttemptSerializer(user_attempts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def highest_attempts(self, request):
        # user_id = request.data.get('user_id', None)
        # user = User.objects.get(id=user_id)
        user = request.user
        attempted_exams = PastExam.objects.filter(exam_attempts__user=user).distinct()
        print("hello world")
        exams_data = []
        for past_exam in attempted_exams:
            attempts = PastExamAttempt.objects.filter(user=user, past_exam=past_exam).annotate(
                position=Window(
                    expression=Rank(),
                    order_by=F('correct_answers').desc()
                )
            )
            highest_attempt = attempts.order_by('-correct_answers').first()
            unique_participants_count = PastExamAttempt.objects.filter(past_exam=past_exam).values('user').distinct().count()

            exams_data.append({
                'past_exam_id': past_exam.id,
                'exam_title': past_exam.title,
                'total_questions': past_exam.total_questions,
                'unique_participants': unique_participants_count,
                'highest_attempt': {
                    'attempt_id': highest_attempt.id,
                    'answered_questions': highest_attempt.answered_questions,
                    'correct_answers': highest_attempt.correct_answers,
                    'wrong_answers': highest_attempt.wrong_answers,
                    'score': highest_attempt.score,
                    'attempt_time': highest_attempt.attempt_time,
                    'position': highest_attempt.position
                }
            })

        return Response({'exams': exams_data}, status=status.HTTP_200_OK)
    
    
    
class CheckPastExamsView(APIView):
    def get(self, request, *args, **kwargs):
        # Get query parameters from the request
        position_id = request.query_params.get('position_id', None)
        department_id = request.query_params.get('department_id', None)
        organization_id = request.query_params.get('organization_id', None)

        # Filtering the past exams based on position, department, or organization
        exams = PastExam.objects.filter(
            position_id=position_id,
            department_id=department_id if department_id else None,
            organization_id=organization_id
        )

        # If no past exams found, return a message indicating no exams available
        if not exams.exists():
            return Response({"message": "No past exams found for this position, department, or organization."},
                            status=status.HTTP_404_NOT_FOUND)

        # If past exams are found, serialize the data
        serializer = PastExamSerializer(exams, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    
class PastExamDeleteView(APIView):
    permission_classes = [IsAuthenticated]  # You can adjust permissions

    def delete(self, request, pk):
        exam = get_object_or_404(PastExam, pk=pk)
        # print(exam)
        # Optional: check if user is allowed to delete (e.g., created_by match)
        if exam.created_by != request.user:
            return Response({'detail': 'Not authorized to delete this exam.'}, status=status.HTTP_403_FORBIDDEN)


        exam.delete()
        return Response({'detail': 'PastExam deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)




class AddQuestionToPastExamView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        past_exam_id = request.data.get('past_exam_id')
        if not past_exam_id:
            return Response({"error": "Missing exam ID"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            exam = PastExam.objects.get(pk=past_exam_id)
        except PastExam.DoesNotExist:
            return Response({"error": "Past exam not found"}, status=status.HTTP_404_NOT_FOUND)
        print(request.data)
        question_text = request.data.get('question_text')
        question_image = request.FILES.get('question_image')
        
        # Set difficulty and question_type to None if not provided
        difficulty = request.data.get('difficulty', None)
        question_type = request.data.get('question_type', None)

        # If neither text nor image is provided, return error
        if not question_text and not question_image:
            return Response({"error": "Question text or image must be provided."}, status=status.HTTP_400_BAD_REQUEST)
        # Parse options (if available)
        options_data = request.data.get('options')
        print(options_data)
        if not options_data:
            return Response({"error": "options are must be provided"})
        # Get or create the question (text-based or image-based)
        if question_text and not question_image:
            question, created = Question.objects.get_or_create(
                text=question_text,
                defaults={
                    'difficulty_level': difficulty,
                    'marks': request.data.get('marks', 1),  # or however you handle marks
                    'time_limit': request.data.get('time_limit', 60)
            }
        )
        else:
            # Create a new image-based question
            question = Question.objects.create(
                text=question_text,
                image=question_image,
                difficulty=difficulty,
                question_type=question_type
            )

        # Create a new PastExamQuestionOption for each option
        past_exam_question = PastExamQuestion.objects.create(
            exam=exam,
            question=question
        )
        
        if options_data:
            try:
                options = json.loads(options_data)
            except json.JSONDecodeError:
                return Response({"error": "Invalid options format."}, status=status.HTTP_400_BAD_REQUEST)

            # Create options for the question (get_or_create to avoid duplicates)
            for idx, opt in enumerate(options):
                if not opt.get('text') and not opt.get('image_key'):
                    return Response({"error": f"Option {idx + 1} must have either text or an image."}, status=status.HTTP_400_BAD_REQUEST)

                # Get or create the option (text or image)
                option, created = QuestionOption.objects.get_or_create(
                    question=question,
                    text=opt.get('text'),
                    image=request.FILES.get(opt.get('image_key')) if opt.get('image_key') else None
                )

                

                # Create the PastExamQuestionOption to link this option to the newly created PastExamQuestion
                PastExamQuestionOption.objects.create(
                    question=past_exam_question,
                    option=option
                )
        

        return Response({"message": "Question with options added to exam."}, status=status.HTTP_201_CREATED)



# Question and Option update view


class UpdateQuestionView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        try:
            question = Question.objects.get(pk=pk)
        except Question.DoesNotExist:
            return Response({"detail": "Question not found."}, status=status.HTTP_404_NOT_FOUND)

        question.text = request.data.get("text", question.text)

        if 'image' in request.FILES:
            uploaded_file = request.FILES['image']
            ext = uploaded_file.name.split('.')[-1]
            unique_filename = f"{uuid.uuid4()}.{ext}"
            file_path = os.path.join("questions", unique_filename)

            saved_path = default_storage.save(file_path, ContentFile(uploaded_file.read()))
            question.image.name = saved_path

        question.save()
        return Response({"detail": "Question updated successfully."})


class UpdateOptionView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        try:
            option = QuestionOption.objects.get(pk=pk)
        except QuestionOption.DoesNotExist:
            return Response({"detail": "Option not found."}, status=status.HTTP_404_NOT_FOUND)

        option.text = request.data.get("text", option.text)

        if 'image' in request.FILES:
            uploaded_file = request.FILES['image']
            ext = uploaded_file.name.split('.')[-1]
            unique_filename = f"{uuid.uuid4()}.{ext}"
            file_path = os.path.join("options", unique_filename)

            saved_path = default_storage.save(file_path, ContentFile(uploaded_file.read()))
            option.image.name = saved_path

        option.save()
        return Response({"detail": "Option updated successfully."})


